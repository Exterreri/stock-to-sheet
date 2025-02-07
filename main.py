from pathlib import Path
import openpyxl.styles
import yfinance as yf
import pandas as pd
import openpyxl

path = Path.cwd()
filename = 'stock_data.xlsx'

# List of stocks I want to get data for
stocks = [
  'AAPL',
  'MSFT',
  'GOOGL',
  'BA',
  'RTX',
  'LMT',
  'NVDA',
  'AMD',
  'INTC',
  'QCOM',
  '^GSPC',
  '^DJI',
]

# Data storage
stock_data = []

# Get data for each stock
for stock in stocks:
  ticker = yf.Ticker(stock)
  info = ticker.info

  # Access balance sheet data
  balance_sheet = ticker.balance_sheet
  financials = ticker.financials

  # Extract data from balance sheet
  total_assets = balance_sheet.loc['Total Assets'].max() if 'Total Assets' in balance_sheet.index else 'N/A'
  total_debt = balance_sheet.loc['Total Debt'].max() if 'Total Debt' in balance_sheet.index else 'N/A'

  # Extract data from financials
  revenue = financials.loc['Total Revenue'].max() if 'Total Revenue' in financials.index else 'N/A'
  gross_profit = financials.loc['Gross Profit'].max() if 'Gross Profit' in financials.index else 'N/A'
  operating_income = financials.loc['Operating Income'].max() if 'Operating Income' in financials.index else 'N/A'
  net_income = financials.loc['Net Income'].max() if 'Net Income' in financials.index else 'N/A'

  # Get the data
  stock_info = {
    'Ticker': stock,  # col A
    'Name': info.get('longName', 'N/A'),  # col B
    'Sector': info.get('sector', 'N/A'),  # col C
    'Industry': info.get('industry', 'N/A'),  # col D
    'Price': info.get('currentPrice', 'N/A'),  # col E
    'Dividend Yield': info.get('dividendRate', 'N/A'),  # col F
    'P/E Ratio': info.get('trailingPE', 'N/A'),  # col G
    'Forward P/E': info.get('forwardPE', 'N/A'),  # col H
    'Market Cap': info.get('marketCap', 'N/A'),  # col I
    'Assets': total_assets,  # col J
    'Debt': total_debt,  # col K
    'Revenue': revenue,  # col L
    'Gross Profit': gross_profit,  # col M
    'Op. Income': operating_income,  # col N
    'Net Income': net_income,  # col O
    'EBITDA': info.get('ebitda', 'N/A'),  # col P
    '(A-L)*1.5': (total_assets - info.get('totalLiab', 0)) * 1.5 if total_assets != 'N/A' else 'N/A',  # col Q
    'Revenue Growth': info.get('revenueGrowth', 'N/A'),  # col R
    'Earnings Growth': info.get('earningsGrowth', 'N/A'),  # col S
    'Current Ratio': info.get('currentRatio', 'N/A'),  # col T
  }
  stock_data.append(stock_info)

# Create a DataFrame
df = pd.DataFrame(stock_data)

# Create a folder under this folder called 'data'
Path(path).parent.mkdir(parents=True, exist_ok=True)

# Save the data to a CSV file
df.to_excel(path / filename, index=False)
print('Data saved to stock_data.xlsx')

# Use openpyxl add formatting to the Excel file
wb = openpyxl.load_workbook(path / filename)
ws = wb.active

currency_columns = [
  ws['I'],  # 'Market Cap'
  ws['J'],  # 'Assets'
  ws['K'],  # 'Debt'
  ws['L'],  # 'Revenue'
  ws['M'],  # 'Gross Profit'
  ws['N'],  # 'Operating Income'
  ws['O'],  # 'Net Income'
  ws['P'],  # 'EBITDA'
  ws['Q'],  # '(Assets - Liabilities) * 1.5'
]

percent_columns = [
  ws['F'],  # 'Dividend Yield'
  ws['R'],  # 'Sales Growth (YoY)'
  ws['S'],  # 'Profit Growth (YoY)'
]

# Apply currency formatting
for col in currency_columns:
  for cell in col[1:]:
    cell.number_format = '$#,##0.00"B"'
    if cell.value != 'N/A':
      cell.value = cell.value / 1e9

# Add dollar signs to the 'Price' column
for cell in ws['E'][1:]:
  cell.number_format = '$#,##0.00'

# Apply percentage formatting
for col in percent_columns:
  for cell in col[1:]:
    # Divide by 100 to get the percentage and add the percentage sign
    cell.number_format = '0.00%'
    if cell.value != 'N/A':
      cell.value = cell.value / 100

# Resize the columns based on the widest data in each column
for col in ws.columns:
  max_length = 0
  column = col[0].column_letter
  for cell in col:
    try:
      if len(str(cell.value)) > max_length:
        max_length = len(str(cell.value))
    except:
      pass
  adjusted_width = (max_length + 2)
  ws.column_dimensions[column].width = adjusted_width

# Color the font in P/E Ratio column
for cell in ws['G'][1:]:
  if cell.value != 'N/A' and cell.value < 15:
    cell.font = openpyxl.styles.Font(color='00FF00')  # Green
  elif cell.value != 'N/A' and cell.value > 25:
    cell.font = openpyxl.styles.Font(color='FF0000')  # Red

# Color the font in Earnings Growth (YoY) column
for cell in ws['S'][1:]:
  if cell.value != 'N/A' and cell.value > 2.9:
    cell.font = openpyxl.styles.Font(color='00FF00')  # Green
  elif cell.value != 'N/A' and cell.value < 0:
    cell.font = openpyxl.styles.Font(color='FF0000')  # Red

# Color the font in Current Ratio column
for cell in ws['T'][1:]:
  if cell.value != 'N/A' and cell.value > 2:
    cell.font = openpyxl.styles.Font(color='00FF00')  # Green
  elif cell.value != 'N/A' and cell.value < 1:
    cell.font = openpyxl.styles.Font(color='FF0000')  # Red

wb.save(path / filename)
print('Formatting applied to stock_data.xlsx')